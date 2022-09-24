from sqlite3 import Row
import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment





keyword = pyautogui.prompt("검색어를 입력하세요 ")
lastpage = int(pyautogui.prompt("몇페이지까지 크롤링 할까요? "))
wb = Workbook()
ws= wb.create_sheet(keyword)

#행번호
row = 1
# 열 너비 조절
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120

pageNum = 1
for i in range(1,lastpage*10,10):
    print(f"================================{pageNum}페이지 크롤링 중입니다================================")
    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&start={i}")
    html = response.text
    soup = BeautifulSoup(html,"html.parser")
    articles= soup.select("div.info_group")

    for article in articles:
        links = article.select("a.info")
        if len(links) >= 2:
            url = links[1].attrs['href']
            response = requests.get(url,headers={"User-Agent":'Mozila/5.0'})
            html = response.text
            soup = BeautifulSoup(html,'html.parser')
            if "entertain" in response.url:
                title = soup.select_one(".end_tit")
                content = soup.select_one("#articeBody")
            # 아니면 스포츠 뉴스라면
            elif "sports" in response.url:
                title = soup.select_one("h4.title")
                content = soup.select_one("#newsEndContents")
                # 본문 내용안에 불필요한 div,p삭제
                divs = content.select('div')
                for div in divs:
                    div.decompose()
                paragraphs = content.select('p')
                for p in paragraphs:
                    p.decompose()
            else:
                title = soup.select_one(".media_end_head_headline")
                content = soup.select_one("#newsct_article")
            print("===========링크===========\n",url)
            print("===========제목===========\n",title.text.strip())
            print("===========본문===========\n",content.text.strip())
            
            ws[f'A{row}']=url
            ws[f'B{row}']=title.text.strip()
            ws[f'C{row}']=content.text.strip()
            
            # 자동 줄바꿈 
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            
            row += 1
            time.sleep(0.3)
    pageNum += 1
    
    
wb.save(f"{keyword}_result.xlsx")
            